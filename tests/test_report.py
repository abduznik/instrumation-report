import datetime
import pytest
from openpyxl import load_workbook

from instrumation_report import Measurement, Report, ReportHeader, Section, TestTable


# ── Helpers ────────────────────────────────────────────────────────────────

def make_report() -> Report:
    report = Report(header=ReportHeader(
        title="RF Subsystem Validation",
        subtitle="Production Test",
        engineer="Yan",
        uut_name="Signal Analyzer Module",
        uut_serial="SN-2024-001",
        revision="A",
    ))
    section = Section(number="1", title="Voltage Tests")
    table = TestTable(title="Power Supply Checks", sub_number="1.1")
    table.add(Measurement("3.3V Rail", 3.31, "V", condition=(3.2, 3.4), test_number="1.1.1", expected="3.2-3.4V"))
    table.add(Measurement("5V Rail", 5.02, "V", condition=(4.9, 5.1), test_number="1.1.2", expected="4.9-5.1V"))
    table.add(Measurement("12V Rail", 11.0, "V", condition=(11.5, 12.5), test_number="1.1.3", expected="11.5-12.5V"))
    table.add(Measurement("Temp", 24.0, "C"))  # N/A
    section.add_table(table)
    report.add_section(section)
    return report


# ── ReportHeader ───────────────────────────────────────────────────────────

def test_header_default_date():
    h = ReportHeader(title="T")
    assert h.date == datetime.date.today().isoformat()

def test_header_date_overridable():
    h = ReportHeader(title="T", date="2020-01-01")
    assert h.date == "2020-01-01"

def test_header_extra_fields():
    h = ReportHeader(title="T", extra_fields={"Batch": "42", "Site": "LA"})
    assert h.extra_fields["Batch"] == "42"


# ── Summary ────────────────────────────────────────────────────────────────

def test_summary_counts():
    r = make_report()
    s = r._build_summary()
    assert s.total == 4
    assert s.passed == 2
    assert s.failed == 1

def test_summary_pass_rate():
    r = make_report()
    s = r._build_summary()
    assert s.pass_rate == "50.0"

def test_summary_empty():
    r = Report()
    s = r._build_summary()
    assert s.total == 0
    assert s.passed == 0
    assert s.failed == 0
    assert s.pass_rate == "0.0"

def test_summary_all_pass():
    r = Report()
    r.add(Measurement("V", 4.5, "V", condition=(4.0, 5.0)))
    r.add(Measurement("F", 1001, "Hz", condition=lambda v: v > 1000))
    s = r._build_summary()
    assert s.passed == 2
    assert s.failed == 0

def test_summary_all_fail():
    r = Report()
    r.add(Measurement("V", 3.0, "V", condition=(4.0, 5.0)))
    r.add(Measurement("F", 500, "Hz", condition=lambda v: v > 1000))
    s = r._build_summary()
    assert s.passed == 0
    assert s.failed == 2


# ── Flat add() backwards compat ────────────────────────────────────────────

def test_flat_add():
    r = Report()
    r.add(Measurement("V", 4.5, "V", condition=(4.0, 5.0)))
    r.add(Measurement("I", 1.2, "A"))
    assert len(r._flat) == 2
    s = r._build_summary()
    assert s.total == 2

def test_flat_and_sections_combined():
    r = Report()
    r.add(Measurement("Flat", 1.0, "V", condition=1.0))
    section = Section(number="1", title="S")
    table = TestTable(title="T", sub_number="1.1")
    table.add(Measurement("Nested", 2.0, "V", condition=(1.5, 2.5)))
    section.add_table(table)
    r.add_section(section)
    assert r._build_summary().total == 2


# ── HTML ───────────────────────────────────────────────────────────────────

def test_html_creates_file(tmp_path):
    out = tmp_path / "report.html"
    make_report().generate_html(str(out))
    assert out.exists()

def test_html_contains_title(tmp_path):
    out = tmp_path / "report.html"
    make_report().generate_html(str(out))
    assert "RF Subsystem Validation" in out.read_text()

def test_html_contains_engineer(tmp_path):
    out = tmp_path / "report.html"
    make_report().generate_html(str(out))
    assert "Yan" in out.read_text()

def test_html_contains_measurements(tmp_path):
    out = tmp_path / "report.html"
    make_report().generate_html(str(out))
    content = out.read_text()
    assert "3.3V Rail" in content
    assert "5V Rail" in content

def test_html_pass_fail_badges(tmp_path):
    out = tmp_path / "report.html"
    make_report().generate_html(str(out))
    content = out.read_text()
    assert "PASS" in content
    assert "FAIL" in content

def test_html_summary(tmp_path):
    out = tmp_path / "report.html"
    make_report().generate_html(str(out))
    content = out.read_text()
    assert "50.0%" in content

def test_html_empty_report(tmp_path):
    out = tmp_path / "empty.html"
    Report().generate_html(str(out))
    content = out.read_text()
    assert out.exists()
    assert "0.0%" in content


# ── Excel ──────────────────────────────────────────────────────────────────

def _get_all_values(ws):
    return [ws.cell(r, c).value for r in range(1, ws.max_row + 1) for c in range(1, ws.max_column + 1)]

def test_excel_creates_file(tmp_path):
    out = tmp_path / "report.xlsx"
    make_report().generate_excel(str(out))
    assert out.exists()

def test_excel_has_col_headers(tmp_path):
    out = tmp_path / "report.xlsx"
    make_report().generate_excel(str(out))
    wb = load_workbook(str(out))
    ws = wb.active
    vals = _get_all_values(ws)
    assert "Test No." in vals
    assert "Pass / Fail" in vals

def test_excel_has_measurement_names(tmp_path):
    out = tmp_path / "report.xlsx"
    make_report().generate_excel(str(out))
    wb = load_workbook(str(out))
    ws = wb.active
    vals = _get_all_values(ws)
    assert "3.3V Rail" in vals
    assert "12V Rail" in vals

def test_excel_status_values(tmp_path):
    out = tmp_path / "report.xlsx"
    make_report().generate_excel(str(out))
    wb = load_workbook(str(out))
    ws = wb.active
    vals = _get_all_values(ws)
    assert "PASS" in vals
    assert "FAIL" in vals
    assert "N/A" in vals

def test_excel_pass_row_fill(tmp_path):
    out = tmp_path / "report.xlsx"
    make_report().generate_excel(str(out))
    wb = load_workbook(str(out))
    ws = wb.active
    fills = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.fgColor:
                fills.add(cell.fill.fgColor.rgb)
    assert any("C6EFCE" in f for f in fills)  # green pass fill
    assert any("FFC7CE" in f for f in fills)  # red fail fill

def test_excel_empty_report(tmp_path):
    out = tmp_path / "empty.xlsx"
    Report().generate_excel(str(out))
    assert out.exists()


# ── PDF ────────────────────────────────────────────────────────────────────

def test_pdf_creates_file(tmp_path):
    pytest.importorskip("weasyprint")
    out = tmp_path / "report.pdf"
    make_report().generate_pdf(str(out))
    assert out.exists()
    assert out.stat().st_size > 0

def test_pdf_no_weasyprint_raises(tmp_path, monkeypatch):
    import sys
    monkeypatch.setitem(sys.modules, "weasyprint", None)
    out = tmp_path / "report.pdf"
    with pytest.raises((ImportError, ModuleNotFoundError)):
        make_report().generate_pdf(str(out))
