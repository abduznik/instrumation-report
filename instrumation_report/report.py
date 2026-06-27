from __future__ import annotations

import datetime
import importlib.resources
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

from jinja2 import BaseLoader, Environment
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .measurement import Measurement
from .section import Section, TestTable


@dataclass
class ReportHeader:
    title: str
    subtitle: str = ""
    engineer: str = ""
    uut_name: str = ""
    uut_serial: str = ""
    revision: str = ""
    date: str = field(default_factory=lambda: datetime.date.today().isoformat())
    extra_fields: dict = field(default_factory=dict)


@dataclass
class _Summary:
    total: int
    passed: int
    failed: int
    pass_rate: str


class Report:
    def __init__(self, header: Optional[ReportHeader] = None) -> None:
        self.header = header or ReportHeader(title="Measurement Report")
        self._sections: List[Section] = []
        self._flat: List[Measurement] = []

    def add_section(self, section: Section) -> None:
        self._sections.append(section)

    def add(self, measurement: Measurement) -> None:
        """Flat backwards-compatible usage — wraps into a default section/table."""
        self._flat.append(measurement)

    def _all_measurements(self) -> List[Measurement]:
        out: List[Measurement] = []
        for s in self._sections:
            for t in s.tables:
                out.extend(t.measurements)
        out.extend(self._flat)
        return out

    def _build_summary(self) -> _Summary:
        all_m = self._all_measurements()
        passed = sum(1 for m in all_m if m.status == "PASS")
        failed = sum(1 for m in all_m if m.status == "FAIL")
        total = len(all_m)
        rate = f"{passed / total * 100:.1f}" if total else "0.0"
        return _Summary(total=total, passed=passed, failed=failed, pass_rate=rate)

    def _sections_for_render(self) -> List[Section]:
        """Merge flat measurements into a synthetic section so the template is uniform."""
        sections = list(self._sections)
        if self._flat:
            s = Section(number="", title="Measurements")
            t = TestTable(title="", sub_number="")
            for m in self._flat:
                t.add(m)
            s.add_table(t)
            sections.append(s)
        return sections

    def _load_template(self) -> str:
        try:
            ref = importlib.resources.files("instrumation_report.templates").joinpath("base.html")
            return ref.read_text(encoding="utf-8")
        except Exception:
            return (Path(__file__).parent / "templates" / "base.html").read_text(encoding="utf-8")

    # ── HTML ──────────────────────────────────────────────────────────────────

    def generate_html(self, output_path: str) -> None:
        env = Environment(loader=BaseLoader())
        tmpl = env.from_string(self._load_template())
        html = tmpl.render(
            header=self.header,
            sections=self._sections_for_render(),
            summary=self._build_summary(),
        )
        Path(output_path).write_text(html, encoding="utf-8")

    # ── Excel ─────────────────────────────────────────────────────────────────

    def generate_excel(self, output_path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        COLS = 7
        COL_LETTERS = [get_column_letter(i) for i in range(1, COLS + 1)]

        # Fills / fonts
        fill_pass    = PatternFill("solid", fgColor="C6EFCE")
        fill_fail    = PatternFill("solid", fgColor="FFC7CE")
        fill_header  = PatternFill("solid", fgColor="1F2937")
        fill_section = PatternFill("solid", fgColor="E5E7EB")
        fill_table   = PatternFill("solid", fgColor="F9FAFB")
        fill_col_hdr = PatternFill("solid", fgColor="F3F4F6")

        font_white   = Font(bold=True, color="FFFFFF", size=11)
        font_section = Font(bold=True, size=11, color="111111")
        font_table   = Font(bold=True, size=10, color="374151")
        font_col_hdr = Font(bold=True, size=9, color="6B7280")
        font_bold    = Font(bold=True)
        font_pass    = Font(color="276221", bold=True)
        font_fail    = Font(color="9B1C1C", bold=True)

        def merge_row(text: str, fill: PatternFill, font: Font, height: float = 18) -> None:
            row = ws.max_row + 1
            ws.append([""] * COLS)
            ws.merge_cells(f"A{row}:{COL_LETTERS[-1]}{row}")
            cell = ws.cell(row, 1)
            cell.value = text
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(vertical="center", indent=1)
            ws.row_dimensions[row].height = height

        def append_col_headers() -> None:
            headers = ["Test No.", "Name", "Value", "Unit", "Expected", "Pass / Fail", "Notes"]
            ws.append(headers)
            r = ws.max_row
            for c, h in enumerate(headers, 1):
                cell = ws.cell(r, c)
                cell.fill = fill_col_hdr
                cell.font = font_col_hdr
                cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[r].height = 16

        # ── Report header block ──
        merge_row(self.header.title, fill_header, font_white, height=28)
        if self.header.subtitle:
            merge_row(self.header.subtitle, fill_header, Font(color="D1D5DB", size=10), height=18)

        meta_pairs = [
            ("Engineer", self.header.engineer),
            ("UUT", self.header.uut_name),
            ("Serial", self.header.uut_serial),
            ("Revision", self.header.revision),
            ("Date", self.header.date),
        ] + list(self.header.extra_fields.items())

        for label, val in meta_pairs:
            if val:
                ws.append([label, val] + [""] * (COLS - 2))
                r = ws.max_row
                ws.cell(r, 1).font = Font(bold=True, color="6B7280", size=9)
                ws.cell(r, 2).font = Font(size=9)

        ws.append([""] * COLS)  # spacer

        # ── Sections ──
        for section in self._sections_for_render():
            if section.number or section.title:
                heading = f"{section.number}. {section.title}".strip(". ")
                merge_row(heading, fill_section, font_section, height=20)
                if section.subtitle:
                    merge_row(section.subtitle, fill_section, Font(italic=True, size=9, color="555555"), height=14)

            for table in section.tables:
                if table.sub_number or table.title:
                    sub_heading = f"{table.sub_number}  {table.title}".strip()
                    merge_row(sub_heading, fill_table, font_table, height=16)

                append_col_headers()

                for m in table.measurements:
                    ws.append([
                        m.test_number,
                        m.name,
                        m.value,
                        m.unit,
                        m.expected,
                        m.status,
                        m.notes,
                    ])
                    r = ws.max_row
                    ws.row_dimensions[r].height = 16
                    status = m.status
                    if status == "PASS":
                        for c in range(1, COLS + 1):
                            ws.cell(r, c).fill = fill_pass
                        ws.cell(r, 6).font = font_pass
                    elif status == "FAIL":
                        for c in range(1, COLS + 1):
                            ws.cell(r, c).fill = fill_fail
                        ws.cell(r, 6).font = font_fail

                ws.append([""] * COLS)  # spacer between tables

        # ── Summary ──
        summary = self._build_summary()
        ws.append([""] * COLS)
        merge_row("Summary", fill_header, font_white, height=20)
        for label, val in [
            ("Total", summary.total),
            ("Passed", summary.passed),
            ("Failed", summary.failed),
            ("Pass Rate", f"{summary.pass_rate}%"),
        ]:
            ws.append([label, val] + [""] * (COLS - 2))
            r = ws.max_row
            ws.cell(r, 1).font = Font(bold=True, size=9, color="6B7280")
            ws.cell(r, 2).font = font_bold

        # Column widths
        widths = [12, 28, 10, 8, 18, 12, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(output_path)

    # ── PDF ───────────────────────────────────────────────────────────────────

    def generate_pdf(self, output_path: str) -> None:
        try:
            from weasyprint import HTML
        except ImportError as e:
            raise ImportError("weasyprint is required for PDF output: pip install weasyprint") from e

        env = Environment(loader=BaseLoader())
        tmpl = env.from_string(self._load_template())
        html = tmpl.render(
            header=self.header,
            sections=self._sections_for_render(),
            summary=self._build_summary(),
        )
        HTML(string=html).write_pdf(output_path)
